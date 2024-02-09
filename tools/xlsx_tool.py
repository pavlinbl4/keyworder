import openpyxl
from pathlib import Path
import string
from openpyxl.styles import (
    Alignment, Font)


#  set columns width in sheet header
def set_column_widths(ws, columns_names, column_width):
    alphabet = string.ascii_uppercase
    for i in range(len(columns_names)):
        ws.column_dimensions[alphabet[i]].width = column_width[i]
        ws.cell(row=1, column=i + 1).font = Font(color="FF0000", size=14, bold=True)
        ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')
    write_column_headers(ws, columns_names)


# Write the column headers if sheet is empty
def write_column_headers(ws, columns_names):
    if ws.max_row == 1:
        for col_num, column in enumerate(columns_names, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column


def write_info_to_xlsx(image_name, destination, path_to_log_folder):
    # Create the file if it doesn't exist yet
    file_path = f"{path_to_log_folder}/uploaded_files.xlsx"
    if not Path(file_path).is_file():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'uploads'
        set_column_widths(ws, columns_names=['File Name',
                                             'Tass-Photo',
                                             'PhotoXpress',
                                             'Kommersant', ],
                          column_width=[50, 50, 50, 50])

        wb.save(file_path)


if __name__ == '__main__':
    write_info_to_xlsx('20230822EPAV3563.JPG',
                       'pavlenko.evgeniy@gmail.com@stringer.photoxpress.ru',
                       '/Users/evgeniy/Library/Mobile Documents/com~apple~CloudDocs/uploads_log')
